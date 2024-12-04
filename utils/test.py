import os
import random
import PyPDF2
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

def extract_text_from_pdf(pdf_path):
    try:
        with open(pdf_path, 'rb') as file:
            reader = PyPDF2.PdfReader(file)
            text = ""
            for page in reader.pages:
                text += page.extract_text() + "\n"
        return text
    except Exception as e:
        return f"Error extracting text: {str(e)}"

def main():
    # Create logs directory if it doesn't exist
    logs_dir = os.path.join(os.path.dirname(__file__), '..', 'logs')
    os.makedirs(logs_dir, exist_ok=True)

    # Create log file with timestamp
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    log_file = os.path.join(logs_dir, f'invoice_text_extraction_{timestamp}.txt')

    # Excel file path
    excel_path = r"C:\Users\JulianMitchell\OneDrive - Cornwells Chemists Limited\Jasper\AI PROGAMMES\INVOICE_PROJECT\Invoice_Summary.xlsx"
    
    # Base path for invoice PDFs
    invoice_base_path = r"C:\Users\JulianMitchell\Cornwells Chemists Limited\Cornwells File Share - Documents\Head Office\Accounts\Invoices\Alliance"
    
    print(f"Starting invoice text extraction...")
    
    with open(log_file, 'w', encoding='utf-8') as log:
        log.write(f"Invoice Text Extraction Log - {datetime.now()}\n")
        log.write("="*80 + "\n\n")

        # Check if paths exist
        if not os.path.exists(excel_path):
            log.write(f"Error: Excel file not found at {excel_path}\n")
            return
            
        if not os.path.exists(invoice_base_path):
            log.write(f"Error: Invoice directory not found at {invoice_base_path}\n")
            return

        # Load the Excel file
        wb = load_workbook(excel_path, read_only=True)
        sheet = wb['Alliance']

        # Get all rows with file paths
        all_rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0]:
                filename = row[0]
                for root, dirs, files in os.walk(invoice_base_path):
                    if os.path.basename(filename) in files:
                        full_path = os.path.join(root, os.path.basename(filename))
                        all_rows.append((full_path, row))
                        break

        print(f"Found {len(all_rows)} invoices in total")
        log.write(f"Found {len(all_rows)} invoices in total\n\n")
        
        # Sample 30 random rows
        sample_size = min(30, len(all_rows))
        sample_rows = random.sample(all_rows, sample_size)
        print(f"Selected {sample_size} random invoices for analysis")
        log.write(f"Selected {sample_size} random invoices for analysis\n\n")

        # Extract and write text from each PDF
        for i, (pdf_path, row) in enumerate(sample_rows, 1):
            try:
                log.write(f"Invoice {i} (File: {os.path.basename(pdf_path)})\n")
                log.write("-"*80 + "\n")
                text = extract_text_from_pdf(pdf_path)
                log.write(text + "\n")
                log.write("-"*80 + "\n\n")
                print(f"Processed invoice {i} of {sample_size}: {os.path.basename(pdf_path)}")
            except Exception as e:
                log.write(f"Error processing {pdf_path}: {str(e)}\n\n")
                print(f"Error processing invoice {i} of {sample_size}")

    print(f"\nExtraction complete! Log file created at: {log_file}")

if __name__ == "__main__":
    main()